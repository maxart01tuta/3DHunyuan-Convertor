"""
Модуль для работы с базой данных Excel.
Обеспечивает получение строк для обработки и обновление статусов.
"""

import openpyxl
from pathlib import Path
from typing import Optional, Dict, Any
from config import EXCEL_FILE

def get_next_row() -> Optional[Dict[str, Any]]:
    """
    Находит первую строку со статусом '#' и возвращает словарь с данными.
    Возвращает None, если строк не найдено.
    """
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    try:
        id_col = header.index('BAZA_ID') + 1
        prompt_col = header.index('BAZA_PROMPT') + 1
        profile_col = header.index('BAZA_PROFILE') + 1
        status_col = header.index('BAZA_GOTOVO') + 1
    except ValueError as e:
        wb.close()
        raise ValueError(f"Отсутствует необходимый столбец в Excel: {e}")

    for row_num in range(2, ws.max_row + 1):
        status = ws.cell(row=row_num, column=status_col).value
        if status == '#':
            row_data = {
                'row_number': row_num,
                'BAZA_ID': ws.cell(row=row_num, column=id_col).value,
                'BAZA_PROMPT': ws.cell(row=row_num, column=prompt_col).value,
                'BAZA_PROFILE': ws.cell(row=row_num, column=profile_col).value,
            }
            wb.close()
            return row_data

    wb.close()
    return None

def mark_row_done(row_number: int):
    """
    Обновляет статус указанной строки на 'ГОТОВО'.
    """
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    status_col = header.index('BAZA_GOTOVO') + 1

    ws.cell(row=row_number, column=status_col).value = 'ГОТОВО'
    wb.save(EXCEL_FILE)
    wb.close()
    print(f"Строка {row_number} обновлена, статус: ГОТОВО")

def mark_row_error(row_number: int):
    """
    Обновляет статус указанной строки на 'ОШИБКА'.
    """
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    status_col = header.index('BAZA_GOTOVO') + 1

    ws.cell(row=row_number, column=status_col).value = 'ОШИБКА'
    wb.save(EXCEL_FILE)
    wb.close()
    print(f"Строка {row_number} обновлена, статус: ОШИБКА")
