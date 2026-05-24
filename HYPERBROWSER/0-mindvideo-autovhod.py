import asyncio
import logging
import os
import sys
import json
import uuid
from pathlib import Path
from typing import Optional, Dict, Any

import openpyxl
from openpyxl import load_workbook

from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError

# Попытка импортировать конфигурацию из той же директории
this_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, this_dir)
import config
import browser
import func_autovhod

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger("mindvideo-autoreg")

# Вспомогательные функции для Excel

def _get_header_indices(ws) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    if ws.max_row < 1:
        return header_map
    first_row = ws[1]
    for cell in first_row:
        if cell.value is None:
            continue
        header_map[str(cell.value).strip().upper()] = cell.column
    return header_map


def _load_workbook() -> load_workbook:
    wb = load_workbook(config.EXCEL_FILE)
    return wb


def get_next_row() -> Optional[Dict[str, Any]]:
    wb = _load_workbook()
    ws = wb["Cookies"]
    headers = _get_header_indices(ws)
    col_gotov = headers.get("BAZA_GOTOVO")
    col_profile = headers.get("BAZA_PROFILE")
    if not col_gotov or not col_profile:
        logger.error("Не найдены необходимые заголовки в Excel: BAZA_GOTOVO или BAZA_PROFILE")
        return None
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=col_gotov).value
        if value is None:
            continue
        if str(value).strip() == "#":
            profile = ws.cell(row=row, column=col_profile).value
            wb.close()
            return {"row_number": row, "BAZA_PROFILE": str(profile).strip() if profile is not None else ""}
    wb.close()
    return None


def mark_row_done(row_number: int) -> None:
    wb = _load_workbook()
    ws = wb["Cookies"]
    headers = _get_header_indices(ws)
    col_gotov = headers.get("BAZA_GOTOVO")
    if not col_gotov:
        logger.error("Не найден столбец BAZA_GOTOVO для обновления статуса ВОШЕЛ")
        wb.close()
        return
    ws.cell(row=row_number, column=col_gotov).value = "ВОШЕЛ"
    wb.save(config.EXCEL_FILE)
    wb.close()


def mark_row_error(row_number: int) -> None:
    wb = _load_workbook()
    ws = wb["Cookies"]
    headers = _get_header_indices(ws)
    col_gotov = headers.get("BAZA_GOTOVO")
    if not col_gotov:
        logger.error("Не найден столбец BAZA_GOTOVO для обновления статуса ОШИБКА")
        wb.close()
        return
    ws.cell(row=row_number, column=col_gotov).value = "ОШИБКА"
    wb.save(config.EXCEL_FILE)
    wb.close()


async def process_row(row_data: Dict[str, Any]) -> bool:
    row_number = int(row_data["row_number"])
    profile_name = str(row_data.get("BAZA_PROFILE", "")).strip()

    playwright = None
    browser_obj = None
    context = None
    page_mindvideo = None
    client = None
    session = None
    released = False

    try:
        auth_path = config.get_auth_path(profile_name)
        session_path = config.get_session_storage_path(profile_name)

        # 1. Запуск Hyperbrowser
        (playwright, browser_obj, client, session) = await browser.launch_hyperbrowser_with_retries(profile_id=profile_name)

        # 2. Настройка контекста (загрузка cookies и sessionStorage)
        context = await browser.setup_context(browser_obj, profile_name)

        # 3. Создание страницы
        page_mindvideo = await context.new_page()

        # 4. Проверка авторизации и автовход (ОДИНАКОВЫЙ ПРОЦЕСС ДЛЯ ВСЕХ)
        await func_autovhod.run(page_mindvideo, profile_name, auth_path, session_path)

        # 5. Обновление Excel
        mark_row_done(row_number)
        logger.info("Строка обновлена как ВОШЕЛ")

        # Этап 7: закрытие браузера
        await browser.release_hyperbrowser(playwright, browser_obj, client, session, context)
        released = True

        return True

    except Exception as e:
        logger.exception(f"Ошибка обработки строки {row_number}: {e}")
        try:
            mark_row_error(row_number)
        except Exception:
            pass
        return False

    finally:
        if not released:
            try:
                await browser.release_hyperbrowser(playwright, browser_obj, client, session, context)
            except Exception:
                pass


async def main() -> None:
    logger.info("Скрипт авторвхода MindVideo запущен")
    while True:
        row = get_next_row()
        if row is None:
            logger.info("Все строки обработаны")
            break
        result = await process_row(row)
        logger.info("ВОШЕЛ" if result else "ОШИБКА")
        await asyncio.sleep(config.WAIT_TIME)


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\nРабота прервана пользователем")
    except Exception as e:
        print(f"Критическая ошибка: {e}")
