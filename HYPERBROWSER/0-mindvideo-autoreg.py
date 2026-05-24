import asyncio
import logging
import os
import sys
import json
import uuid
from pathlib import Path
from typing import Optional, Dict, Any

import openpyxl
from openpyxl import load_workbook

from playwright.async_api import async_playwright

# Попытка импортировать конфигурацию из той же директории
this_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, this_dir)
import config

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger("mindvideo-autoreg")

# Вспомогательные функции для Excel

def _get_header_indices(ws) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    if ws.max_row < 1:
        return header_map
    first_row = ws[1]
    for cell in first_row:
        if cell.value is None:
            continue
        header_map[str(cell.value).strip().upper()] = cell.column
    return header_map


def _load_workbook() -> load_workbook:
    wb = load_workbook(config.EXCEL_FILE)
    return wb


def get_next_row() -> Optional[Dict[str, Any]]:
    wb = _load_workbook()
    ws = wb["Cookies"]
    headers = _get_header_indices(ws)
    col_gotov = headers.get("BAZA_GOTOVO")
    col_profile = headers.get("BAZA_PROFILE")
    if not col_gotov or not col_profile:
        logger.error("Не найдены необходимые заголовки в Excel: BAZA_GOTOVO или BAZA_PROFILE")
        return None
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=col_gotov).value
        if value is None:
            continue
        if str(value).strip() == "#":
            profile = ws.cell(row=row, column=col_profile).value
            wb.close()
            return {"row_number": row, "BAZA_PROFILE": str(profile).strip() if profile is not None else ""}
    wb.close()
    return None


def mark_row_done(row_number: int) -> None:
    wb = _load_workbook()
    ws = wb["Cookies"]
    headers = _get_header_indices(ws)
    col_gotov = headers.get("BAZA_GOTOVO")
    if not col_gotov:
        logger.error("Не найден столбец BAZA_GOTOVO для обновления статуса ГОТОВО")
        wb.close()
        return
    ws.cell(row=row_number, column=col_gotov).value = "ГОТОВО"
    wb.save(config.EXCEL_FILE)
    wb.close()


def mark_row_error(row_number: int) -> None:
    wb = _load_workbook()
    ws = wb["Cookies"]
    headers = _get_header_indices(ws)
    col_gotov = headers.get("BAZA_GOTOVO")
    if not col_gotov:
        logger.error("Не найден столбец BAZA_GOTOVO для обновления статуса ОШИБКА")
        wb.close()
        return
    ws.cell(row=row_number, column=col_gotov).value = "ОШИБКА"
    wb.save(config.EXCEL_FILE)
    wb.close()


async def launch_browser():
    # импорт внутри функции, чтобы не загружать до необходимости
    from hyperbrowser import Hyperbrowser
    from hyperbrowser.models import CreateSessionParams

    api_keys = list(config.API_HYPERBROWSER_LIST)
    import random
    random.shuffle(api_keys)

    last_exc: Optional[Exception] = None

    for attempt in range(max(1, config.MAX_HYPERBROWSER_RETRIES)):
        api_key = api_keys[attempt % len(api_keys)]
        client = Hyperbrowser(api_key=api_key)
        try:
            session = client.sessions.create(
                CreateSessionParams(
                    accept_cookies=True,
                    screen={"width": 1920, "height": 1080},
                    save_downloads=True,
                )
            )
            # Подключение через CDP к Playwright
            playwright = await async_playwright().start()
            browser = await playwright.chromium.connect_over_cdp(session.ws_endpoint)
            context = browser.contexts[0] if browser.contexts else await browser.new_context()
            page = context.pages[0] if context.pages else await context.new_page()
            # Настройки таймаутов хардкодом через config
            # Возвращаем коллекцию объектов, чтобы можно было управлять ими извне
            return (playwright, browser, context, page, client, session)
        except Exception as e:
            last_exc = e
            logger.exception(f"Ошибка при создании сессии Hyperbrowser (попытка {attempt+1}): {e}")
            try:
                if 'client' in locals() and client is not None:
                    try:
                        if 'session' in locals() and session is not None:
                            client.sessions.stop(session.id)
                    except Exception:
                        pass
                if 'playwright' in locals() and playwright is not None:
                    try:
                        await playwright.stop()
                    except Exception:
                        pass
            except Exception:
                pass
            await asyncio.sleep(config.HYPERBROWSER_RETRY_DELAY)
            continue
    raise RuntimeError("Не удалось запустить Hyperbrowser после нескольких попыток")


async def release_browser(playwright, browser, client, session_id) -> None:
    try:
        if browser:
            try:
                await browser.close()
            except Exception:
                pass
        if client and session_id:
            try:
                client.sessions.stop(session_id)
            except Exception:
                pass
        if playwright:
            try:
                await playwright.stop()
            except Exception:
                pass
    except Exception:
        logger.exception("Ошибка при завершении работы браузера Hyperbrowser")


async def save_profile(context, baza_profile: str) -> str:
    base_dir = config.PROFILES_FOLDER
    profile_dir = os.path.join(base_dir, str(baza_profile))
    Path(profile_dir).mkdir(parents=True, exist_ok=True)
    auth_path = os.path.join(profile_dir, config.AUTH_FILENAME)
    await context.storage_state(path=auth_path, indexed_db=True)
    logger.info(f"Профиль сохранён: {auth_path}")
    return auth_path


async def process_row(row_data: Dict[str, Any]) -> bool:
    row_number = int(row_data["row_number"])
    profile_name = str(row_data.get("BAZA_PROFILE", "")).strip()

    playwright = None
    browser = None
    context = None
    page_mindvideo = None
    client = None
    session = None

    try:
        # Этап 1: запустить браузер БЕЗ профиля (первая регистрация)
        (playwright, browser, context, page_mindvideo, client, session) = await launch_browser()
        logger.info(f"Браузер запущен, сессия ID={getattr(session, 'id', None)}")

        # Этап 2: регистрация на MindVideo на первой вкладке
        await page_mindvideo.goto(config.site_registration, timeout=config.MAX_TIME * 1000, wait_until='domcontentloaded')
        await asyncio.sleep(config.WAIT_TIME)

        # Email
        await page_mindvideo.wait_for_selector(config.input_mail_registration, timeout=config.MAX_TIME * 1000)
        await page_mindvideo.fill(config.input_mail_registration, f"b-{profile_name}@merepost.com")

        # Nickname
        await page_mindvideo.wait_for_selector(config.input_nickname_registration, timeout=config.MAX_TIME * 1000)
        await page_mindvideo.fill(config.input_nickname_registration, f"b-{profile_name}merepost")

        # Password
        await page_mindvideo.wait_for_selector(config.input_password_registration, timeout=config.MAX_TIME * 1000)
        await page_mindvideo.fill(config.input_password_registration, config.mindvideo_password)

        await asyncio.sleep(config.WAIT_TIME)

        # Continue
        await page_mindvideo.wait_for_selector(config.knopka_continue_registration, timeout=config.MAX_TIME * 1000)
        await page_mindvideo.click(config.knopka_continue_registration)
        logger.info("Возможно в ручную нажать Cloudflare")
        await asyncio.sleep(config.WAIT_TIME)

        # Ожидание кода верификации (первая вкладка)
        await page_mindvideo.wait_for_selector(config.input_code_verify, state='visible', timeout=config.MAX_TIME * 1000)
        logger.info("Форма верификации появилась, открываем tempmail.plus")

        # Этап 3: получение кода на второй вкладке (tempmail)
        page_tempmail = await context.new_page()
        try:
            await page_tempmail.goto(config.site_merepost, timeout=config.MAX_TIME * 1000, wait_until='domcontentloaded')
            await asyncio.sleep(config.WAIT_TIME)

            await page_tempmail.wait_for_selector(config.input_mail_merepost, timeout=config.MAX_TIME * 1000)
            await page_tempmail.fill(config.input_mail_merepost, f"b-{profile_name}")
            await asyncio.sleep(config.WAIT_TIME)

            await page_tempmail.wait_for_selector(config.knopka_dropdown_mails, timeout=config.MAX_TIME * 1000)
            await page_tempmail.click(config.knopka_dropdown_mails)
            await asyncio.sleep(config.WAIT_TIME)

            await page_tempmail.wait_for_selector(config.option_merepost, timeout=config.MAX_TIME * 1000)
            await page_tempmail.click(config.option_merepost)
            await asyncio.sleep(config.WAIT_TIME)

            await page_tempmail.wait_for_selector(config.text_verify_mindvideo, state='visible', timeout=config.MAX_TIME * 1000)
            await page_tempmail.click(config.text_verify_mindvideo)
            await asyncio.sleep(config.WAIT_TIME)

            await page_tempmail.wait_for_selector(config.text_verify_code, state='visible', timeout=config.MAX_TIME * 1000)
            code_element = await page_tempmail.wait_for_selector(config.text_verify_code, state='visible', timeout=config.MAX_TIME * 1000)
            await code_element.scroll_into_view_if_needed()
            code = await code_element.text_content()
            code = code.strip() if code else ""
            if not code:
                raise ValueError("Не получен код верификации из письма")
            logger.info(f"Получен код: {code}")
        finally:
            try:
                await page_tempmail.close()
            except Exception:
                pass

        # Этап 4: Ввод кода на MindVideo (первая вкладка)
        await page_mindvideo.wait_for_selector(config.input_code_verify, timeout=config.MAX_TIME * 1000)
        await page_mindvideo.fill(config.input_code_verify, code)
        await asyncio.sleep(config.WAIT_TIME)

        await page_mindvideo.wait_for_selector(config.knopka_verify_registration, timeout=config.MAX_TIME * 1000)
        await page_mindvideo.click(config.knopka_verify_registration)
        await asyncio.sleep(config.WAIT_TIME)

        await page_mindvideo.wait_for_selector(config.text_account_mindvideo, state='visible', timeout=config.MAX_TIME * 1000)
        logger.info("Авторизация успешна")
        await asyncio.sleep(config.WAIT_TIME)
        await page_mindvideo.reload(timeout=config.MAX_TIME * 1000, wait_until='domcontentloaded')
        await asyncio.sleep(config.WAIT_TIME)
        await page_mindvideo.wait_for_selector(config.text_account_mindvideo, state='visible', timeout=config.MAX_TIME * 1000)
        await asyncio.sleep(config.WAIT_TIME)

        # Этап 5: сохранение профиля
        await save_profile(context, profile_name)
        logger.info("Профиль сохранён")

        # Этап 6: закрытие браузера и обновление Excel
        await release_browser(playwright, browser, client, getattr(session, 'id', None))
        mark_row_done(row_number)
        logger.info("Строка обновлена как ГОТОВО")
        return True
    except Exception as e:
        logger.exception(f"Ошибка обработки строки {row_number}: {e}")
        try:
            mark_row_error(row_number)
        except Exception:
            pass
        return False
    finally:
        try:
            await release_browser(playwright, browser, client, getattr(session, 'id', None))
        except Exception:
            pass


async def main() -> None:
    logger.info("Скрипт авторегистрации MindVideo запущен")
    while True:
        row = get_next_row()
        if row is None:
            logger.info("Все строки обработаны")
            break
        result = await process_row(row)
        logger.info("ГОТОВО" if result else "ОШИБКА")
        await asyncio.sleep(config.WAIT_TIME)


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\nРабота прервана пользователем")
    except Exception as e:
        print(f"Критическая ошибка: {e}")
