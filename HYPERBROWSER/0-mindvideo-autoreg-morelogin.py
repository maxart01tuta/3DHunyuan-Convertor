import asyncio
import logging
import os
import sys
import json
import uuid
import random
from pathlib import Path
from typing import Optional, Dict, Any, Tuple
from urllib.parse import urlparse

import aiohttp
import openpyxl
from openpyxl import load_workbook

from playwright.async_api import async_playwright

# Попытка импортировать конфигурацию из той же директории
this_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, this_dir)
import config

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger("mindvideo-autoreg")

# Вспомогательные функции для Excel

def _get_header_indices(ws) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    if ws.max_row < 1:
        return header_map
    first_row = ws[1]
    for cell in first_row:
        if cell.value is None:
            continue
        header_map[str(cell.value).strip().upper()] = cell.column
    return header_map


def _load_workbook() -> load_workbook:
    wb = load_workbook(config.EXCEL_FILE)
    return wb


def get_next_row() -> Optional[Dict[str, Any]]:
    wb = _load_workbook()
    ws = wb["Cookies"]
    headers = _get_header_indices(ws)
    col_gotov = headers.get("BAZA_GOTOVO")
    col_profile = headers.get("BAZA_PROFILE")
    if not col_gotov or not col_profile:
        logger.error("Не найдены необходимые заголовки в Excel: BAZA_GOTOVO или BAZA_PROFILE")
        return None
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=col_gotov).value
        if value is None:
            continue
        if str(value).strip() == "#":
            profile = ws.cell(row=row, column=col_profile).value
            wb.close()
            return {"row_number": row, "BAZA_PROFILE": str(profile).strip() if profile is not None else ""}
    wb.close()
    return None


def mark_row_done(row_number: int) -> None:
    wb = _load_workbook()
    ws = wb["Cookies"]
    headers = _get_header_indices(ws)
    col_gotov = headers.get("BAZA_GOTOVO")
    if not col_gotov:
        logger.error("Не найден столбец BAZA_GOTOVO для обновления статуса ЗАРЕГЕСТРИРОВАН")
        wb.close()
        return
    ws.cell(row=row_number, column=col_gotov).value = "ЗАРЕГЕСТРИРОВАН"
    wb.save(config.EXCEL_FILE)
    wb.close()


def mark_row_error(row_number: int) -> None:
    wb = _load_workbook()
    ws = wb["Cookies"]
    headers = _get_header_indices(ws)
    col_gotov = headers.get("BAZA_GOTOVO")
    if not col_gotov:
        logger.error("Не найден столбец BAZA_GOTOVO для обновления статуса ОШИБКА")
        wb.close()
        return
    ws.cell(row=row_number, column=col_gotov).value = "ОШИБКА"
    wb.save(config.EXCEL_FILE)
    wb.close()


async def launch_browser(row_profile_name: str):
    """
    Берет фиксированный MoreLogin профиль из конфига, обновляет прокси и подключает Playwright.
    """
    token = choose_random_morelogin_token()
    proxy_url = choose_random_proxy_url()
    proxy_data = parse_proxy_url_to_morelogin_proxy(proxy_url)
    
    client = MoreLoginClient(token)
    
    env_id = config.MORELOGIN_PROFILE_ID
    playwright = None
    browser = None
    
    try:
        logger.info(f"Используем фиксированный профиль env_id={env_id}")

        # 1. Обновление прокси в профиле
        proxy_user = proxy_data.get("username", "unknown")
        logger.info(f"Обновление прокси для профиля {env_id} (User: {proxy_user}, IP: {proxy_data.get('proxyIp')})...")
        await client.set_proxy(env_id, proxy_data)
        
        # 2. Обновление фингерпринта
        await client.refresh_fingerprint(env_id)
        
        # 3. Запуск профиля
        logger.info(f"Запуск профиля env_id={env_id}")
        debug_port = await client.start_env(env_id)
        
        # 5. Подключение Playwright через CDP
        playwright = await async_playwright().start()
        
        # Ретраи для подключения к CDP
        last_exc = None
        for attempt in range(5):
            try:
                browser = await playwright.chromium.connect_over_cdp(f"http://127.0.0.1:{debug_port}")
                break
            except Exception as e:
                last_exc = e
                logger.warning(f"Попытка подключения к CDP ({attempt+1}/5) не удалась: {e}")
                await asyncio.sleep(2)
        else:
            raise RuntimeError(f"Не удалось подключиться к CDP на порту {debug_port} после 5 попыток: {last_exc}")
            
        context = browser.contexts[0] if browser.contexts else await browser.new_context()
        page = context.pages[0] if context.pages else await context.new_page()
        
        return (playwright, browser, context, page, client, env_id)
        
    except Exception as e:
        logger.error(f"Ошибка при запуске браузера MoreLogin: {e}")
        if playwright:
            await playwright.stop()
        raise


async def release_browser(playwright, browser, client, env_id) -> None:
    """
    Закрывает браузер и очищает кэш MoreLogin (БЕЗ удаления профиля).
    """
    try:
        if browser:
            try:
                # ПРИНУДИТЕЛЬНАЯ ОЧИСТКА ЧЕРЕЗ PLAYWRIGHT ПЕРЕД ЗАКРЫТИЕМ
                for context in browser.contexts:
                    try:
                        await context.clear_cookies()
                        # Очистка localStorage/sessionStorage на всех открытых страницах
                        for page in context.pages:
                            try:
                                await page.evaluate("localStorage.clear(); sessionStorage.clear();")
                            except:
                                pass
                    except:
                        pass
                await browser.close()
            except Exception:
                pass
        
        if client and env_id:
            try:
                logger.info(f"Закрытие и очистка профиля MoreLogin env_id={env_id}")
                await client.close_env(env_id)
                
                # Небольшая пауза, чтобы MoreLogin успел синхронизировать закрытие
                await asyncio.sleep(3)
                
                await client.remove_local_cache(env_id)
                await client.refresh_fingerprint(env_id)
                # Удаление профиля (delete_env) исключено для стратегии Reuse
            except Exception as e:
                logger.warning(f"Ошибка при очистке MoreLogin env_id={env_id}: {e}")
                
        if playwright:
            try:
                await playwright.stop()
            except Exception:
                pass
    except Exception as e:
        logger.exception(f"Критическая ошибка при завершении работы браузера: {e}")


# --- MoreLogin API Utilities ---

def choose_random_morelogin_token() -> Optional[str]:
    tokens = config.API_MORELOGIN_LIST
    if not tokens or (len(tokens) == 1 and not tokens[0]):
        # API_MORELOGIN_LIST отключен или пуст
        return None
    return random.choice(tokens)


def choose_random_proxy_url() -> str:
    proxies = config.PROXIES
    if not proxies:
        raise ValueError("config.PROXIES пуст")
    return random.choice(proxies)


def parse_proxy_url_to_morelogin_proxy(proxy_url: str) -> Dict[str, Any]:
    """
    Парсит URL прокси в строгом соответствии с документацией MoreLogin Local API:
    - proxyProvider: string ("0": http, "1": https, "2": socks5)
    - proxyType: integer (0: http, 1: https)
    """
    parsed = urlparse(proxy_url)
    scheme = parsed.scheme.upper()
    
    # proxyProvider (string): 0: http, 1: https, 2: socks5
    # proxyType (int): 0: http, 1: https (для socks5 обычно тоже 0 или игнорируется)
    if "SOCKS" in scheme:
        provider = "2"
        p_type = 0
    elif "HTTPS" in scheme:
        provider = "1"
        p_type = 1
    else:
        provider = "0"
        p_type = 0
        
    try:
        port = int(parsed.port) if parsed.port else (443 if provider == "1" else 80)
    except:
        port = 80

    return {
        "proxyProvider": provider,
        "proxyType": p_type,
        "proxyIp": (parsed.hostname or "").strip(),
        "proxyPort": port,
        "username": (parsed.username or "").strip(),
        "password": (parsed.password or "").strip(),
        # Дополнительные поля, которые могут требоваться валидатором
        "proxyMethod": 2,
        "ipMonitor": False
    }


# --- MoreLogin API Utilities (Aliases for plan compatibility) ---

async def morelogin_create_env_quick(client: MoreLoginClient, proxy_data: Dict[str, Any], env_name: str) -> str:
    return await client.create_env_quick(proxy_data, env_name)

async def morelogin_refresh_fingerprint(client: MoreLoginClient, env_id: str) -> bool:
    return await client.refresh_fingerprint(env_id)

async def morelogin_start_env(client: MoreLoginClient, env_id: str) -> int:
    return await client.start_env(env_id)

async def morelogin_close_env(client: MoreLoginClient, env_id: str) -> bool:
    return await client.close_env(env_id)

async def morelogin_remove_local_cache(client: MoreLoginClient, env_id: str) -> bool:
    return await client.remove_local_cache(env_id)

async def morelogin_delete_env(client: MoreLoginClient, env_id: str) -> bool:
    return await client.delete_env(env_id)


class MoreLoginClient:
    def __init__(self, api_token: Optional[str] = None):
        self.base_url = config.MORELOGIN_BASE_URL.rstrip('/')
        self.headers = {
            "Content-Type": "application/json"
        }
        if api_token:
            self.headers["Authorization"] = api_token

    async def _request(self, method: str, endpoint: str, json_data: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        url = f"{self.base_url}{endpoint}"
        last_exc = None
        
        # Если токен не задан, не добавляем заголовок Authorization
        current_headers = self.headers.copy()
        if not current_headers.get("Authorization"):
            current_headers.pop("Authorization", None)
            # logger.debug(f"Запрос без авторизации: {endpoint}")
        
        for attempt in range(config.MAX_MORELOGIN_RETRIES):
            try:
                # Используем один сеанс для запроса
                async with aiohttp.ClientSession() as session:
                    kwargs = {
                        "headers": current_headers,
                        "timeout": aiohttp.ClientTimeout(total=config.MORELOGIN_HTTP_TIMEOUT)
                    }
                    if json_data is not None:
                        kwargs["json"] = json_data
                    
                    async with session.request(method, url, **kwargs) as response:
                        response_text = await response.text()
                        try:
                            result = json.loads(response_text)
                        except json.JSONDecodeError:
                            logger.error(f"MoreLogin API вернул не JSON: {response_text}")
                            raise RuntimeError(f"Invalid JSON response: {response_text}")

                        if response.status == 200:
                            if isinstance(result, list):
                                return {"code": 0, "data": result}
                            if result.get("code") == 0 or result.get("success") is True:
                                return result
                        
                        logger.warning(f"MoreLogin API error ({endpoint}): {result}. Payload sent: {json_data}. Попытка {attempt+1}")
            except Exception as e:
                last_exc = e
                logger.warning(f"Ошибка запроса MoreLogin ({endpoint}): {e}. Попытка {attempt+1}")
            
            await asyncio.sleep(config.MORELOGIN_RETRY_DELAY)
            
        raise RuntimeError(f"Не удалось выполнить запрос MoreLogin {endpoint} после {config.MAX_MORELOGIN_RETRIES} попыток: {last_exc}")

    async def create_env_quick(self, proxy_data: Dict[str, Any], env_name: str) -> str:
        payload = {
            "name": env_name,
            "browserTypeId": 1,  # 1 - Chrome, 2 - Firefox
            "operatorSystemId": 1,  # 1 - Windows, 2 - macOS, 3 - Linux, 4 - Android, 5 - iOS
            "quantity": 1,
            "proxy": proxy_data
        }
        result = await self._request("POST", "/api/env/create/quick", payload)
        data = result.get("data")
        
        # Если в data пришел список, берем первый элемент
        if isinstance(data, list) and len(data) > 0:
            env_id = data[0].get("envId") or data[0].get("id")
        elif isinstance(data, dict):
            env_id = data.get("envId") or data.get("id")
        else:
            env_id = None
            
        if not env_id:
            raise ValueError(f"Не удалось получить envId из ответа create/quick: {result}")
        return str(env_id)

    async def refresh_fingerprint(self, env_id: str) -> bool:
        payload = {"envId": env_id}
        try:
            result = await self._request("POST", "/api/env/fingerprint/refresh", payload)
            return result.get("code") == 0 or result.get("success") is True
        except Exception as e:
            logger.warning(f"Не удалось обновить фингерпринт для {env_id}: {e}")
            return False

    async def start_env(self, env_id: str) -> int:
        payload = {"envId": env_id}
        # Start может занимать время
        result = await self._request("POST", "/api/env/start", payload)
        debug_port = result.get("data", {}).get("debugPort")
        if not debug_port:
            raise ValueError(f"Не удалось получить debugPort из ответа start: {result}")
        return int(debug_port)

    async def close_env(self, env_id: str) -> bool:
        payload = {"envId": env_id}
        try:
            result = await self._request("POST", "/api/env/close", payload)
            return result.get("code") == 0 or result.get("success") is True
        except Exception as e:
            logger.warning(f"Ошибка при закрытии профиля {env_id}: {e}")
            return False

    async def remove_local_cache(self, env_id: str) -> bool:
        local_payload = {
            "envId": str(env_id),
            "cookie": True,
            "localStorage": True,
            "indexedDB": True,
            "extension": True,
            "extensionFile": True
        }
        cloud_payload = {
            "envId": str(env_id),
            "cookie": True,
            "others": True
        }
        try:
            # 1. Очистка локального кэша (файлы на диске)
            await self._request("POST", "/api/env/removeLocalCache", local_payload)
            
            # 2. Очистка облачного кэша (чтобы данные не восстановились из синхронизации)
            # Правильный путь согласно документации: /api/env/cache/cleanCloud
            await self._request("POST", "/api/env/cache/cleanCloud", cloud_payload)
            
            return True
        except Exception as e:
            logger.warning(f"Ошибка при полной очистке кэша (Local/Cloud) {env_id}: {e}")
            return False

    async def delete_env(self, env_id: str) -> bool:
        payload = {"envIds": [env_id]}
        try:
            result = await self._request("POST", "/api/env/removeToRecycleBin/batch", payload)
            return result.get("code") == 0 or result.get("success") is True
        except Exception as e:
            logger.warning(f"Ошибка при удалении профиля {env_id}: {e}")
            return False

    async def set_proxy(self, env_id: str, proxy_data: Dict[str, Any]) -> bool:
        # envIds должен быть списком строк, так как ID профилей слишком длинные для int32
        payload = {
            "envIds": [str(env_id)],
            "proxy": proxy_data
        }
        try:
            # Используем специализированный эндпоинт для пакетного обновления прокси
            # Он более надежен для изменения настроек прокси в существующих профилях
            result = await self._request("POST", "/api/env/setProxy/batch", payload)
            return result.get("code") == 0 or result.get("success") is True
        except Exception as e:
            logger.warning(f"Ошибка при установке прокси для {env_id}: {e}")
            return False

    async def get_envs(self, page_no: int = 1, page_size: int = 50) -> list:
        payload = {"pageNo": page_no, "pageSize": page_size}
        try:
            result = await self._request("POST", "/api/env/page", payload)
            return result.get("data", {}).get("list", [])
        except Exception as e:
            logger.warning(f"Ошибка при получении списка профилей: {e}")
            return []


async def process_row(row_data: Dict[str, Any]) -> bool:
    row_number = int(row_data["row_number"])
    profile_name = str(row_data.get("BAZA_PROFILE", "")).strip()

    playwright = None
    browser = None
    context = None
    page_mindvideo = None
    client = None
    env_id = None

    try:
        # Этап 1: запустить браузер MoreLogin
        (playwright, browser, context, page_mindvideo, client, env_id) = await launch_browser(profile_name)
        logger.info(f"Браузер запущен, env_id={env_id}")

        # Этап 2: регистрация на MindVideo на первой вкладке
        await page_mindvideo.goto(config.site_registration, timeout=config.MAX_TIME * 1000, wait_until='domcontentloaded')
        await asyncio.sleep(config.WAIT_TIME)

        # Email
        await page_mindvideo.wait_for_selector(config.input_mail_registration, timeout=config.MAX_TIME * 1000)
        await page_mindvideo.fill(config.input_mail_registration, f"b-{profile_name}@merepost.com")

        # Nickname
        await page_mindvideo.wait_for_selector(config.input_nickname_registration, timeout=config.MAX_TIME * 1000)
        await page_mindvideo.fill(config.input_nickname_registration, f"b-{profile_name}merepost")

        # Password
        await page_mindvideo.wait_for_selector(config.input_password_registration, timeout=config.MAX_TIME * 1000)
        await page_mindvideo.fill(config.input_password_registration, config.mindvideo_password)

        await asyncio.sleep(config.WAIT_TIME)

        # Continue
        await page_mindvideo.wait_for_selector(config.knopka_continue_registration, timeout=config.MAX_TIME * 1000)
        await page_mindvideo.click(config.knopka_continue_registration)
        logger.info("Возможно в ручную нажать Cloudflare")
        await asyncio.sleep(config.WAIT_TIME)

        # Ожидание кода верификации (первая вкладка)
        await page_mindvideo.wait_for_selector(config.input_code_verify, state='visible', timeout=config.MAX_TIME * 1000)
        logger.info("Форма верификации появилась, открываем tempmail.plus")

        # Этап 3: получение кода на второй вкладке (tempmail)
        page_tempmail = await context.new_page()
        try:
            await page_tempmail.goto(config.site_merepost, timeout=config.MAX_TIME * 1000, wait_until='domcontentloaded')
            await asyncio.sleep(config.WAIT_TIME)

            await page_tempmail.wait_for_selector(config.input_mail_merepost, timeout=config.MAX_TIME * 1000)
            await page_tempmail.fill(config.input_mail_merepost, f"b-{profile_name}")
            await asyncio.sleep(config.WAIT_TIME)

            await page_tempmail.wait_for_selector(config.knopka_dropdown_mails, timeout=config.MAX_TIME * 1000)
            await page_tempmail.click(config.knopka_dropdown_mails)
            await asyncio.sleep(config.WAIT_TIME)

            await page_tempmail.wait_for_selector(config.option_merepost, timeout=config.MAX_TIME * 1000)
            await page_tempmail.click(config.option_merepost)
            await asyncio.sleep(config.WAIT_TIME)

            await page_tempmail.wait_for_selector(config.text_verify_mindvideo, state='visible', timeout=config.MAX_TIME * 1000)
            await page_tempmail.click(config.text_verify_mindvideo)
            await asyncio.sleep(config.WAIT_TIME)

            await page_tempmail.wait_for_selector(config.text_verify_code, state='visible', timeout=config.MAX_TIME * 1000)
            code_element = await page_tempmail.wait_for_selector(config.text_verify_code, state='visible', timeout=config.MAX_TIME * 1000)
            await code_element.scroll_into_view_if_needed()
            code = await code_element.text_content()
            code = code.strip() if code else ""
            if not code:
                raise ValueError("Не получен код верификации из письма")
            logger.info(f"Получен код: {code}")
        finally:
            try:
                await page_tempmail.close()
            except Exception:
                pass

        # Этап 4: Ввод кода на MindVideo (первая вкладка)
        await page_mindvideo.wait_for_selector(config.input_code_verify, timeout=config.MAX_TIME * 1000)
        await page_mindvideo.fill(config.input_code_verify, code)
        await asyncio.sleep(config.WAIT_TIME)

        await page_mindvideo.wait_for_selector(config.knopka_verify_registration, timeout=config.MAX_TIME * 1000)
        await page_mindvideo.click(config.knopka_verify_registration)
        await asyncio.sleep(config.WAIT_TIME)

        await page_mindvideo.wait_for_selector(config.text_account_mindvideo, state='visible', timeout=config.MAX_TIME * 1000)
        logger.info("Авторизация успешна")
        await asyncio.sleep(config.WAIT_TIME)
        # переагрузка страницы не нужна await page_mindvideo.reload(timeout=config.MAX_TIME * 1000, wait_until='domcontentloaded')
        # переагрузка страницы не нужна await asyncio.sleep(config.WAIT_TIME)
        # переагрузка страницы не нужна await page_mindvideo.wait_for_selector(config.text_account_mindvideo, state='visible', timeout=config.MAX_TIME * 1000)
        # переагрузка страницы не нужна await asyncio.sleep(config.WAIT_TIME)

        # Этап 5: обновление Excel
        mark_row_done(row_number)
        logger.info("Строка обновлена как ЗАРЕГЕСТРИРОВАН")
        return True
    except Exception as e:
        logger.exception(f"Ошибка обработки строки {row_number}: {e}")
        try:
            mark_row_error(row_number)
        except Exception:
            pass
        return False
    finally:
        try:
            await release_browser(playwright, browser, client, env_id)
        except Exception:
            pass


async def main() -> None:
    logger.info("Скрипт авторегистрации MindVideo запущен")
    while True:
        row = get_next_row()
        if row is None:
            logger.info("Все строки обработаны")
            break
        result = await process_row(row)
        logger.info("ГОТОВО" if result else "ОШИБКА")
        await asyncio.sleep(config.WAIT_TIME)


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\nРабота прервана пользователем")
    except Exception as e:
        print(f"Критическая ошибка: {e}")
