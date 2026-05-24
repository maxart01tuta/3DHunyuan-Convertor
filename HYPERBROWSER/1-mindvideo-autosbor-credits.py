import asyncio
import json
import logging
import os
import random
import sys
from pathlib import Path
from typing import Optional, Dict, Any

from openpyxl import load_workbook
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError

# Импорт config.py из текущей директории
THIS_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, THIS_DIR)
import config
import browser
import func_autovhod


# =========================
# ЛОГИРОВАНИЕ
# =========================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger("mindvideo-autosbor-credits")


# =========================
# EXCEL: ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# =========================
def _get_header_indices(ws) -> Dict[str, int]:
    """
    Возвращает словарь вида:
    {
        "BAZA_PROFILE": 1,
        "BAZA_GOTOVO": 2,
        ...
    }
    """
    header_map: Dict[str, int] = {}
    if ws.max_row < 1:
        return header_map

    for cell in ws[1]:
        if cell.value is None:
            continue
        header_map[str(cell.value).strip().upper()] = cell.column

    return header_map


def _load_workbook():
    return load_workbook(config.EXCEL_FILE_2)


def get_next_row() -> Optional[Dict[str, Any]]:
    """
    Находит первую строку на листе Cookies, где BAZA_GOTOVO == '#'
    и возвращает:
    {
        "row_number": 2,
        "BAZA_PROFILE": "01"
    }
    """
    wb = _load_workbook()
    try:
        ws = wb["Cookies"]
        headers = _get_header_indices(ws)

        col_gotovo = headers.get("BAZA_GOTOVO")
        col_profile = headers.get("BAZA_PROFILE")

        if not col_gotovo or not col_profile:
            logger.error("Не найдены заголовки BAZA_GOTOVO или BAZA_PROFILE в Excel.")
            return None

        for row in range(2, ws.max_row + 1):
            value = ws.cell(row=row, column=col_gotovo).value
            if value is None:
                continue

            if str(value).strip() == "#":
                profile = ws.cell(row=row, column=col_profile).value
                return {
                    "row_number": row,
                    "BAZA_PROFILE": str(profile).strip() if profile is not None else "",
                }

        return None
    finally:
        wb.close()


def set_row_status(row_number: int, status: str) -> None:
    """
    Универсальная запись статуса в BAZA_GOTOVO.
    """
    wb = _load_workbook()
    try:
        ws = wb["Cookies"]
        headers = _get_header_indices(ws)
        col_gotovo = headers.get("BAZA_GOTOVO")

        if not col_gotovo:
            logger.error("Не найден столбец BAZA_GOTOVO для обновления статуса.")
            return

        ws.cell(row=row_number, column=col_gotovo).value = status
        wb.save(config.EXCEL_FILE_2)
    finally:
        wb.close()


def mark_row_done(row_number: int, tokens: str) -> None:
    set_row_status(row_number, f"ТОКЕНОВ: {tokens}")


def mark_row_error(row_number: int) -> None:
    set_row_status(row_number, "ОШИБКА")


# =========================
# BROWSER HELPERS
# =========================
async def wait_short_pause():
    await asyncio.sleep(config.WAIT_TIME)


async def is_check_in_button_disabled(locator) -> bool:
    """
    Определяем disabled максимально надежно.
    """
    try:
        if await locator.is_disabled():
            return True
    except Exception:
        pass

    try:
        attr_disabled = await locator.get_attribute("disabled")
        if attr_disabled is not None:
            return True
    except Exception:
        pass

    try:
        aria_disabled = await locator.get_attribute("aria-disabled")
        if aria_disabled and aria_disabled.lower() == "true":
            return True
    except Exception:
        pass

    try:
        class_name = await locator.get_attribute("class")
        if class_name and "disabled" in class_name.lower():
            return True
    except Exception:
        pass

    return False


async def wait_until_button_becomes_disabled(page, locator):
    """
    После клика ждем, пока кнопка станет disabled.
    """
    handle = await locator.element_handle()
    if handle is None:
        raise RuntimeError("Не удалось получить element_handle для кнопки Check In.")

    await page.wait_for_function(
        """
        (el) => {
            if (!el) return false;
            const byDisabledAttr = el.hasAttribute('disabled');
            const byAria = (el.getAttribute('aria-disabled') || '').toLowerCase() === 'true';
            const byClass = (el.className || '').toString().toLowerCase().includes('disabled');
            return byDisabledAttr || byAria || byClass;
        }
        """,
        arg=handle,
        timeout=config.MAX_TIME * 1000,
    )


async def process_row(row_data: Dict[str, Any]) -> bool:
    row_number = int(row_data["row_number"])
    profile_name = str(row_data.get("BAZA_PROFILE", "")).strip()

    logger.info(f"Начата обработка строки Excel #{row_number}, профиль: {profile_name}")

    if not profile_name:
        logger.error(f"Пустой BAZA_PROFILE в строке #{row_number}")
        mark_row_error(row_number)
        return False

    auth_path = config.get_auth_path(profile_name)
    session_path = config.get_session_storage_path(profile_name)
    logger.info(f"Путь к auth state: {auth_path}")

    playwright = None
    browser_obj = None
    client = None
    session = None
    context = None

    try:
        # 1. Запуск Hyperbrowser
        playwright, browser_obj, client, session = await browser.launch_hyperbrowser_with_retries(profile_id=profile_name)

        # 2. Настройка контекста (загрузка cookies и sessionStorage)
        context = await browser.setup_context(browser_obj, profile_name)

        # 3. Создание страницы
        page = await context.new_page()

        # 4. Проверка авторизации и автовход (ОДИНАКОВЫЙ ПРОЦЕСС ДЛЯ ВСЕХ)
        await func_autovhod.run(page, profile_name, auth_path, session_path)

        # 5. Переход на страницу text-to-image (хотя func_autovhod уже мог там быть)
        if page.url != config.site_text_to_image:
            await page.goto(config.site_text_to_image, wait_until="domcontentloaded", timeout=config.MAX_TIME * 1000)
            logger.info(f"Открыта страница: {config.site_text_to_image}")
            await wait_short_pause()

        # 6. Клик по 'Check In & Claim'
        claim_locator = page.locator(config.knopka_check_in_claim).first
        await claim_locator.wait_for(state="visible", timeout=config.MAX_TIME * 1000)
        await claim_locator.click()
        logger.info("Открыл Check In and Claim")
        await asyncio.sleep(config.WAIT_TIME * 2)

# 7. Работа с кнопкой 'Check In'
        check_in_locator = page.locator(config.knopka_check_in).first
        check_in_disabled_locator = page.locator(config.knopka_check_in_disabled).first

        if await check_in_locator.count() > 0:
            await check_in_locator.click()
            logger.info("Кредиты успешно собраны")
            await asyncio.sleep(config.WAIT_TIME)
        elif await check_in_disabled_locator.count() > 0:
            logger.info("Новых кредитов нет")
            await asyncio.sleep(config.WAIT_TIME)
        else:
            raise RuntimeError("Не найдена ни активная, ни disabled кнопка Check In")

        # 8. Получаем количество токенов
        tokens = "???"
        try:
            tokens_locator = page.locator(config.text_tokens_count).first
            await tokens_locator.wait_for(state="visible", timeout=10000)
            tokens = await tokens_locator.inner_text()
            tokens = tokens.strip()
            logger.info(f"Количество токенов на сайте: {tokens}")
        except Exception as e:
            logger.warning(f"Не удалось получить количество токенов: {e}")

        logger.info(f"Строка #{row_number} успешно завершена. ТОКЕНОВ: {tokens}")
        mark_row_done(row_number, tokens)
        return True

    except PlaywrightTimeoutError as e:
        logger.error(f"Timeout при обработке строки #{row_number}: {e}")
        mark_row_error(row_number)
        return False

    except (FileNotFoundError, json.JSONDecodeError, ValueError) as e:
        logger.error(f"Ошибка auth state для строки #{row_number}: {e}")
        mark_row_error(row_number)
        return False

    except Exception as e:
        logger.exception(f"Ошибка при обработке строки #{row_number}: {e}")
        mark_row_error(row_number)
        return False

    finally:
        try:
            await browser.release_hyperbrowser(
                playwright=playwright,
                browser=browser_obj,
                client=client,
                session=session,
                context=context,
            )
        except Exception:
            pass


# =========================
# MAIN LOOP
# =========================
async def main() -> None:
    logger.info("Скрипт автосбора кредитов MindVideo запущен.")

    while True:
        row = get_next_row()
        if row is None:
            logger.info("Строк со статусом '#' больше не найдено. Работа завершена.")
            break

        result = await process_row(row)
        logger.info("Итог обработки: СОБРАНО" if result else "Итог обработки: ОШИБКА")

        # Короткая пауза перед следующей строкой
        await wait_short_pause()

    logger.info("Скрипт завершил работу.")


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\nРабота прервана пользователем.")
    except Exception as e:
        print(f"Критическая ошибка: {e}")