import openpyxl
import os
from typing import Optional, Dict, Any
import config

def get_next_row() -> Optional[Dict[str, Any]]:
    """
    Находит первую строку со статусом '#' в листе 'MindVideo' и возвращает словарь с данными.
    """
    if not os.path.exists(config.EXCEL_FILE):
        print(f"Файл Excel не найден: {config.EXCEL_FILE}")
        return None

    wb = openpyxl.load_workbook(config.EXCEL_FILE)
    if "MindVideo" not in wb.sheetnames:
        wb.close()
        print(f"Лист 'MindVideo' не найден в {config.EXCEL_FILE}")
        return None
        
    ws = wb["MindVideo"]

    header = [cell.value for cell in ws[1]]
    try:
        id_col = header.index('BAZA_ID') + 1
        url_col = header.index('BAZA_URL') + 1
        promt_col = header.index('BAZA_PROMT') + 1
        profile_col = header.index('BAZA_PROFILE') + 1
        status_col = header.index('BAZA_GOTOVO') + 1
    except ValueError as e:
        wb.close()
        print(f"Отсутствует необходимый столбец в Excel (Лист MindVideo): {e}")
        return None

    for row_num in range(2, ws.max_row + 1):
        status = ws.cell(row=row_num, column=status_col).value
        if status == '#':
            row_data = {
                'row_number': row_num,
                'BAZA_ID': ws.cell(row=row_num, column=id_col).value,
                'BAZA_URL': ws.cell(row=row_num, column=url_col).value,
                'BAZA_PROMT': ws.cell(row=row_num, column=promt_col).value,
                'BAZA_PROFILE': ws.cell(row=row_num, column=profile_col).value,
            }
            wb.close()
            return row_data

    wb.close()
    return None

def mark_row_status(row_number: int, status: str, tokens: Optional[str] = None):
    """
    Обновляет статус строки и опционально записывает количество токенов.
    """
    if not os.path.exists(config.EXCEL_FILE):
        return

    wb = openpyxl.load_workbook(config.EXCEL_FILE)
    ws = wb["MindVideo"]

    header = [cell.value for cell in ws[1]]
    status_col = header.index('BAZA_GOTOVO') + 1
    
    # Записываем статус
    ws.cell(row=row_number, column=status_col).value = status
    
    # Если переданы токены и есть столбец для них
    if tokens is not None:
        try:
            tokens_col = header.index('BAZA_TOKENS') + 1
            ws.cell(row=row_number, column=tokens_col).value = tokens
        except ValueError:
            # Если столбца BAZA_TOKENS нет, можно записывать в статус или лог
            pass

    wb.save(config.EXCEL_FILE)
    wb.close()
    print(f"Строка {row_number} обновлена, статус: {status}" + (f", токены: {tokens}" if tokens else ""))
